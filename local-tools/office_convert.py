#!/usr/bin/env python3
"""
Office 文件转换工具 - AiBoxDash 本地工具集
支持: CSV↔Excel、Markdown→HTML、JSON↔CSV、文本编码转换、PDF 文本提取

用法:
  python office_convert.py --action <csv2xlsx|xlsx2csv|md2html|json2csv|csv2json|txt2pdf|pdf2txt|merge_csv|split_csv> --input <文件> [--output <输出>] [--sheet 0] [--encoding utf-8]

输出: JSON 格式（转换结果、输出路径）
依赖:
  - 基础功能（csv/json/html）: 仅标准库
  - Excel 读写: pip install openpyxl
  - PDF 操作: pip install fpdf2 PyPDF2
"""
import argparse
import csv
import html
import json
import os
import re
import sys
from pathlib import Path


def csv_to_xlsx(input_path: str, output_path: str = None, sheet_name: str = "Sheet1") -> dict:
    """CSV 转 Excel（需要 openpyxl）"""
    try:
        from openpyxl import Workbook
    except ImportError:
        return {"ok": False, "error": "需要安装 openpyxl: pip install openpyxl"}

    inp = Path(input_path)
    out = Path(output_path) if output_path else inp.with_suffix(".xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    with open(inp, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    wb.save(str(out))
    return {"ok": True, "output": str(out.resolve()), "rows": ws.max_row, "cols": ws.max_column}


def xlsx_to_csv(input_path: str, output_path: str = None, sheet_index: int = 0) -> dict:
    """Excel 转 CSV（需要 openpyxl）"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"ok": False, "error": "需要安装 openpyxl: pip install openpyxl"}

    inp = Path(input_path)
    out = Path(output_path) if output_path else inp.with_suffix(".csv")

    wb = load_workbook(str(inp), read_only=True, data_only=True)
    sheets = wb.sheetnames
    if sheet_index >= len(sheets):
        return {"ok": False, "error": f"工作表索引 {sheet_index} 超出范围（共 {len(sheets)} 个）"}

    ws = wb[sheets[sheet_index]]
    with open(out, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([str(cell) if cell is not None else "" for cell in row])

    wb.close()
    return {"ok": True, "output": str(out.resolve()), "sheet": sheets[sheet_index]}


def md_to_html(input_path: str, output_path: str = None) -> dict:
    """Markdown 转 HTML（简易转换，支持标题/列表/代码块/粗体/链接）"""
    inp = Path(input_path)
    out = Path(output_path) if output_path else inp.with_suffix(".html")

    text = inp.read_text(encoding="utf-8")
    lines = text.split("\n")
    html_lines = []
    in_code = False
    in_list = False

    for line in lines:
        # 代码块
        if line.strip().startswith("```"):
            if in_code:
                html_lines.append("</code></pre>")
                in_code = False
            else:
                lang = line.strip()[3:].strip()
                html_lines.append(f'<pre><code class="language-{lang}">')
                in_code = True
            continue
        if in_code:
            html_lines.append(html.escape(line))
            continue

        # 标题
        m = re.match(r"^(#{1,6})\s+(.+)$", line)
        if m:
            level = len(m.group(1))
            html_lines.append(f"<h{level}>{m.group(2)}</h{level}>")
            continue

        # 无序列表
        if re.match(r"^[-*+]\s+", line):
            if not in_list:
                html_lines.append("<ul>")
                in_list = True
            html_lines.append(f"<li>{line[2:]}</li>")
            continue
        elif in_list:
            html_lines.append("</ul>")
            in_list = False

        # 空行
        if not line.strip():
            html_lines.append("")
            continue

        # 普通段落（内联格式）
        line = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", line)
        line = re.sub(r"\*(.+?)\*", r"<em>\1</em>", line)
        line = re.sub(r"\[(.+?)\]\((.+?)\)", r'<a href="\2">\1</a>', line)
        line = re.sub(r"`(.+?)`", r"<code>\1</code>", line)
        html_lines.append(f"<p>{line}</p>")

    if in_list:
        html_lines.append("</ul>")

    body = "\n".join(html_lines)
    full_html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="utf-8"><title>{inp.stem}</title>
<style>body{{font-family:system-ui;max-width:800px;margin:2em auto;padding:0 1em;line-height:1.6}}
pre{{background:#f5f5f5;padding:1em;overflow-x:auto}}code{{font-family:monospace}}</style>
</head><body>{body}</body></html>"""

    out.write_text(full_html, encoding="utf-8")
    return {"ok": True, "output": str(out.resolve()), "size_bytes": out.stat().st_size}


def json_to_csv(input_path: str, output_path: str = None) -> dict:
    """JSON 数组转 CSV"""
    inp = Path(input_path)
    out = Path(output_path) if output_path else inp.with_suffix(".csv")

    data = json.loads(inp.read_text(encoding="utf-8"))
    if not isinstance(data, list) or len(data) == 0:
        return {"ok": False, "error": "JSON 必须是非空数组"}

    # 收集所有键
    keys = []
    for item in data:
        if isinstance(item, dict):
            for k in item.keys():
                if k not in keys:
                    keys.append(k)

    with open(out, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
        writer.writeheader()
        for item in data:
            if isinstance(item, dict):
                writer.writerow({k: str(v) if v is not None else "" for k, v in item.items()})

    return {"ok": True, "output": str(out.resolve()), "rows": len(data), "columns": keys}


def csv_to_json(input_path: str, output_path: str = None) -> dict:
    """CSV 转 JSON 数组"""
    inp = Path(input_path)
    out = Path(output_path) if output_path else inp.with_suffix(".json")

    with open(inp, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    out.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "output": str(out.resolve()), "rows": len(rows)}


def merge_csv(input_path: str, output_path: str = None) -> dict:
    """合并目录下所有 CSV 文件（input_path 为目录）"""
    dir_path = Path(input_path)
    if not dir_path.is_dir():
        return {"ok": False, "error": "输入路径必须是目录"}

    csv_files = sorted(dir_path.glob("*.csv"))
    if not csv_files:
        return {"ok": False, "error": "目录下没有 CSV 文件"}

    out = Path(output_path) if output_path else dir_path / "_merged.csv"
    all_rows = []
    headers = None

    for f in csv_files:
        with open(f, "r", encoding="utf-8-sig") as fp:
            reader = csv.DictReader(fp)
            if headers is None:
                headers = reader.fieldnames
            all_rows.extend(reader)

    with open(out, "w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=headers or [])
        writer.writeheader()
        writer.writerows(all_rows)

    return {"ok": True, "output": str(out.resolve()), "files_merged": len(csv_files), "total_rows": len(all_rows)}


def pdf_to_txt(input_path: str, output_path: str = None) -> dict:
    """PDF 文本提取（需要 PyPDF2）"""
    try:
        from PyPDF2 import PdfReader
    except ImportError:
        return {"ok": False, "error": "需要安装 PyPDF2: pip install PyPDF2"}

    inp = Path(input_path)
    out = Path(output_path) if output_path else inp.with_suffix(".txt")

    reader = PdfReader(str(inp))
    text_parts = []
    for page in reader.pages:
        text_parts.append(page.extract_text() or "")

    full_text = "\n\n".join(text_parts)
    out.write_text(full_text, encoding="utf-8")
    return {"ok": True, "output": str(out.resolve()), "pages": len(reader.pages), "chars": len(full_text)}


def txt_to_pdf(input_path: str, output_path: str = None) -> dict:
    """文本转 PDF（需要 fpdf2）"""
    try:
        from fpdf import FPDF
    except ImportError:
        return {"ok": False, "error": "需要安装 fpdf2: pip install fpdf2"}

    inp = Path(input_path)
    out = Path(output_path) if output_path else inp.with_suffix(".pdf")

    pdf = FPDF()
    pdf.add_page()
    # 使用内置字体（中文需要额外字体文件，此处用 latin-1 兼容）
    pdf.set_font("Helvetica", size=11)

    text = inp.read_text(encoding="utf-8")
    for line in text.split("\n"):
        # fpdf 不支持直接输出中文，使用 latin-1 安全编码
        safe_line = line.encode("latin-1", errors="replace").decode("latin-1")
        pdf.cell(0, 6, safe_line, new_x="LMARGIN", new_y="NEXT")

    pdf.output(str(out))
    return {"ok": True, "output": str(out.resolve()), "size_bytes": out.stat().st_size}


def main():
    parser = argparse.ArgumentParser(description="Office 文件转换工具")
    parser.add_argument("--action", required=True,
                        choices=["csv2xlsx", "xlsx2csv", "md2html", "json2csv", "csv2json",
                                 "merge_csv", "pdf2txt", "txt2pdf"],
                        help="转换操作")
    parser.add_argument("--input", required=True, help="输入文件/目录路径")
    parser.add_argument("--output", default=None, help="输出文件路径（默认自动推断）")
    parser.add_argument("--sheet", type=int, default=0, help="Excel 工作表索引（默认 0）")
    args = parser.parse_args()

    actions = {
        "csv2xlsx": lambda: csv_to_xlsx(args.input, args.output),
        "xlsx2csv": lambda: xlsx_to_csv(args.input, args.output, args.sheet),
        "md2html": lambda: md_to_html(args.input, args.output),
        "json2csv": lambda: json_to_csv(args.input, args.output),
        "csv2json": lambda: csv_to_json(args.input, args.output),
        "merge_csv": lambda: merge_csv(args.input, args.output),
        "pdf2txt": lambda: pdf_to_txt(args.input, args.output),
        "txt2pdf": lambda: txt_to_pdf(args.input, args.output),
    }

    result = actions[args.action]()
    print(json.dumps(result, ensure_ascii=False, indent=2))
    sys.exit(0 if result.get("ok") else 1)


if __name__ == "__main__":
    main()
